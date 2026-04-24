from __future__ import annotations

import csv
from io import BytesIO, StringIO

from openpyxl import load_workbook
from app.services.importers.base import normalize_key


KNOWN_HEADER_CELLS = {
    "data",
    "dataksiegowania",
    "dataoperacji",
    "opis",
    "opisoperacji",
    "tytul",
    "nadawcaodbiorca",
    "numerkonta",
    "kwota",
    "saldopooperacji",
    "numerfaktury",
    "kontrahent",
    "kwotabrutto",
    "formaplatnosci",
    "typdokumentu",
    "rodzajdokumentu",
}


def read_tabular_bytes(filename: str, payload: bytes) -> list[dict[str, object]]:
    suffix = filename.lower().rsplit(".", 1)[-1]
    if suffix == "csv":
        return _read_csv(payload)
    if suffix in {"xlsx", "xlsm"}:
        return _read_xlsx(payload)
    raise ValueError(f"Unsupported file format: {filename}")


def _read_csv(payload: bytes) -> list[dict[str, object]]:
    text = _decode_csv(payload)
    sample = text[:2048]
    delimiter = ";" if sample.count(";") >= sample.count(",") else ","
    lines = text.splitlines()
    header_row = _find_header_row(lines, delimiter)
    reader = csv.DictReader(StringIO("\n".join(lines[header_row:])), delimiter=delimiter)
    return [dict(row) for row in reader]


def _decode_csv(payload: bytes) -> str:
    for encoding in ("utf-8-sig", "cp1250", "utf-8", "latin-1"):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    return payload.decode("utf-8", errors="ignore")


def _find_header_row(lines: list[str], delimiter: str) -> int:
    best_index = 0
    best_score = -1

    for index, line in enumerate(lines):
        if not line.strip():
            continue
        cells = next(csv.reader([line], delimiter=delimiter))
        normalized_cells = [normalize_key(cell.strip()) for cell in cells if cell.strip()]
        if len(normalized_cells) < 2:
            continue
        score = sum(1 for cell in normalized_cells if cell in KNOWN_HEADER_CELLS)
        if score > best_score:
            best_index = index
            best_score = score

    return best_index


def _read_xlsx(payload: bytes) -> list[dict[str, object]]:
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    result: list[dict[str, object]] = []
    for row in rows[1:]:
        result.append({headers[index]: value for index, value in enumerate(row)})
    return result
